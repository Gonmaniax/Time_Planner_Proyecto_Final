from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.models.database import db
from app.models.tarea import Tarea
from app.models.sesion_cronometro import SesionCronometro
from app.models.categoria import Categoria
from datetime import datetime, timedelta, date
import io

reportes_bp = Blueprint("reportes", __name__)

def formato_tiempo(minutos):
    """Convierte minutos a formato legible: 1h 30min"""
    if minutos == 0:
        return "0 min"
    horas = minutos // 60
    mins = minutos % 60
    if horas > 0 and mins > 0:
        return f"{horas}h {mins}min"
    elif horas > 0:
        return f"{horas}h"
    else:
        return f"{mins}min"

def _to_date(valor):
    """Normaliza datetime o date a date, sin romper si el campo es None."""
    if valor is None:
        return None
    return valor.date() if hasattr(valor, 'date') else valor

def get_datos_reporte(usuario_id):
    hoy = date.today()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    fin_semana = inicio_semana + timedelta(days=6)

    # Todas las tareas del usuario (histórico completo, se filtra después)
    tareas = Tarea.query.filter_by(id_usuario=usuario_id).all()
    tareas_ids = [t.id for t in tareas]

    # Sesiones de esta semana
    sesiones = SesionCronometro.query.filter(
        SesionCronometro.id_tarea.in_(tareas_ids),
        db.func.date(SesionCronometro.inicio) >= inicio_semana,
        db.func.date(SesionCronometro.inicio) <= fin_semana
    ).all() if tareas_ids else []

    total_min = sum(s.duracion_min for s in sesiones)

    # Tiempo real por tarea, calculado desde las sesiones reales de esta semana
    # (en vez de confiar en un campo que podría no actualizarse siempre)
    tiempo_real_por_tarea = {}
    for s in sesiones:
        tiempo_real_por_tarea[s.id_tarea] = tiempo_real_por_tarea.get(s.id_tarea, 0) + s.duracion_min

    def pertenece_a_la_semana(t):
        """
        Regla única de pertenencia a la semana del reporte, usada de forma
        consistente en TODAS las métricas para que los números siempre cuadren:
        - Completada -> pertenece a la semana en que se completó.
        - Pendiente/en progreso -> pertenece si su fecha límite cae esta semana,
          o si no tiene fecha límite pero se creó esta semana.
        """
        if t.estado == 'completada':
            fc = _to_date(t.fecha_completada)
            return fc is not None and inicio_semana <= fc <= fin_semana
        fl = _to_date(t.fecha_limite)
        if fl is not None:
            return inicio_semana <= fl <= fin_semana
        fcreacion = _to_date(t.fecha_creacion)
        return fcreacion is not None and inicio_semana <= fcreacion <= fin_semana

    tareas_semana = [t for t in tareas if pertenece_a_la_semana(t)]

    completadas = [t for t in tareas_semana if t.estado == 'completada']
    pendientes = [t for t in tareas_semana if t.estado == 'pendiente']
    

    tiempo_estimado_total = sum(t.tiempo_estimado_min for t in completadas)
    tiempo_real_total = sum(tiempo_real_por_tarea.get(t.id, 0) for t in completadas)

    # Por categoría (usa el mismo conjunto tareas_semana, así los números
    # de aquí siempre suman lo mismo que las métricas generales de arriba)
    por_categoria = []
    categorias = Categoria.query.filter(
        (Categoria.id_usuario == usuario_id) |
        (Categoria.tipo == 'predeterminada')
    ).all()

    for cat in categorias:
        tareas_cat_semana = [t for t in tareas_semana if t.id_categoria == cat.id]
        ids_cat = [t.id for t in tareas_cat_semana]
        min_cat = sum(tiempo_real_por_tarea.get(tid, 0) for tid in ids_cat)
        completadas_cat = len([t for t in tareas_cat_semana if t.estado == 'completada'])
        total_cat = len(tareas_cat_semana)
        if min_cat > 0 or total_cat > 0:
            por_categoria.append({
                "nombre": cat.nombre,
                "minutos": min_cat,
                "tiempo_formato": formato_tiempo(min_cat),
                "porcentaje": round((min_cat / total_min * 100) if total_min > 0 else 0),
                "completadas": completadas_cat,
                "total": total_cat
            })

    # Por día de la semana (sin cambios, ya estaba correcto)
    dias = []
    nombres_dias = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    for i in range(7):
        dia = inicio_semana + timedelta(days=i)
        sesiones_dia = [s for s in sesiones if s.inicio.date() == dia]
        min_dia = sum(s.duracion_min for s in sesiones_dia)
        dias.append({
            "nombre": nombres_dias[i],
            "fecha": dia.strftime('%d/%m'),
            "minutos": min_dia,
            "tiempo_formato": formato_tiempo(min_dia),
            "es_hoy": dia == hoy
        })

    # Detalle de TODAS las tareas de la semana (completadas, pendientes y en progreso)
    # para que se vea con claridad qué se hizo, qué falta, y con qué tiempo real
    orden_estado = {'completada': 0, 'en_progreso': 1, 'pendiente': 2}
    detalle_tareas_semana = []
    for t in tareas_semana:
        tiempo_real_min = tiempo_real_por_tarea.get(t.id, 0)
        con_cronometro = tiempo_real_min > 0

        if t.estado == 'completada':
            if con_cronometro:
                tiempo_real_txt = formato_tiempo(tiempo_real_min)
                diferencia = tiempo_real_min - t.tiempo_estimado_min
            else:
                tiempo_real_txt = 'Completada sin cronómetro'
                diferencia = None
        else:
            tiempo_real_txt = formato_tiempo(tiempo_real_min) if con_cronometro else '—'
            diferencia = None

        detalle_tareas_semana.append({
            'titulo': t.titulo,
            'estado': t.estado,
            'con_cronometro': con_cronometro,
            'tiempo_estimado': formato_tiempo(t.tiempo_estimado_min),
            'tiempo_real': tiempo_real_txt,
            'diferencia': diferencia
        })

    detalle_tareas_semana.sort(key=lambda x: orden_estado.get(x['estado'], 3))

    return {
        "total_tareas": len(tareas_semana),
        "completadas": len(completadas),
        "pendientes": len(pendientes),
        "tiempo_total": formato_tiempo(total_min),
        "tiempo_total_min": total_min,
        "tiempo_estimado_total": formato_tiempo(tiempo_estimado_total),
        "tiempo_real_total": formato_tiempo(tiempo_real_total),
        "por_categoria": sorted(por_categoria, key=lambda x: x['minutos'], reverse=True),
        "por_dia": dias,
        "detalle_tareas_semana": detalle_tareas_semana,
        "inicio_semana": inicio_semana.strftime('%d/%m/%Y'),
        "fin_semana": fin_semana.strftime('%d/%m/%Y'),
        "max_minutos_dia": max((d['minutos'] for d in dias), default=1) or 1
    }

@reportes_bp.route("/semana", methods=["GET"])
@jwt_required()
def get_reporte_semana():
    usuario_id = int(get_jwt_identity())
    datos = get_datos_reporte(usuario_id)
    return jsonify(datos), 200

@reportes_bp.route("/exportar", methods=["GET"])
@jwt_required()
def exportar_reporte():
    usuario_id = int(get_jwt_identity())
    formato = request.args.get("formato", "pdf")
    datos = get_datos_reporte(usuario_id)
    if formato == "pdf":
        return exportar_pdf(datos)
    elif formato == "excel":
        return exportar_excel(datos)
    return jsonify({"error": "Formato no soportado"}), 400

def exportar_pdf(datos):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    elements = []

    CAFE     = colors.HexColor('#3D1F0D')
    TERRACOTA= colors.HexColor('#C0623A')
    CREMA    = colors.HexColor('#FDF6EC')
    BORDE    = colors.HexColor('#E8D5B0')
    GRIS     = colors.HexColor('#8B5A2B')

    s_titulo   = ParagraphStyle('t', fontSize=22, leading=26, fontName='Helvetica-Bold', textColor=CAFE, spaceAfter=10)
    s_sub      = ParagraphStyle('s', fontSize=11, leading=14, fontName='Helvetica', textColor=GRIS, spaceAfter=16)
    s_seccion  = ParagraphStyle('sec', fontSize=13, fontName='Helvetica-Bold', textColor=CAFE, spaceBefore=14, spaceAfter=6)

    elements.append(Paragraph("Time Planner", s_titulo))
    elements.append(Paragraph(f"Reporte de productividad — Semana del {datos['inicio_semana']} al {datos['fin_semana']}", s_sub))
    elements.append(Spacer(1, 0.2*cm))

    # Métricas
    elements.append(Paragraph("Resumen general", s_seccion))
    met = [
    ["Tareas totales", "Completadas", "Pendientes", "Tiempo trabajado"],
    [str(datos['total_tareas']), str(datos['completadas']),
     str(datos['pendientes']), datos['tiempo_total']]
]
    t_met = Table(met, colWidths=[3.8*cm]*4)
    t_met.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), CAFE),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('BACKGROUND', (0,1), (-1,1), CREMA),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,1), (-1,1), 13),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, BORDE),
        ('ROWHEIGHT', (0,0), (-1,-1), 28),
    ]))
    elements.append(t_met)
    elements.append(Spacer(1, 0.3*cm))

    # Por categoría
    if datos['por_categoria']:
        elements.append(Paragraph("Tiempo por categoría", s_seccion))
        cat_h = [["Categoría", "Tiempo", "% del total", "Tareas completadas"]]
        for cat in datos['por_categoria']:
            cat_h.append([cat['nombre'], cat['tiempo_formato'],
                         f"{cat['porcentaje']}%", str(cat['completadas'])])
        t_cat = Table(cat_h, colWidths=[5*cm, 3.5*cm, 3.5*cm, 4*cm])
        t_cat.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), TERRACOTA),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [CREMA, colors.white]),
            ('GRID', (0,0), (-1,-1), 0.5, BORDE),
            ('PADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(t_cat)

    # Detalle de tareas de la semana (completadas, pendientes y en progreso)
    if datos['detalle_tareas_semana']:
        elements.append(Paragraph("Detalle de tareas de la semana", s_seccion))
        etiquetas_estado = {'completada': 'Completada', 'pendiente': 'Pendiente', 'en_progreso': 'En progreso'}
        tar_h = [["Tarea", "Estado", "Estimado", "Real"]]
        for t in datos['detalle_tareas_semana']:
            tar_h.append([
                t['titulo'],
                etiquetas_estado.get(t['estado'], t['estado']),
                t['tiempo_estimado'],
                t['tiempo_real']
            ])
        t_tar = Table(tar_h, colWidths=[6.5*cm, 3*cm, 3*cm, 3.5*cm])
        t_tar.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), CAFE),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [CREMA, colors.white]),
            ('GRID', (0,0), (-1,-1), 0.5, BORDE),
            ('PADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(t_tar)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True,
                     download_name='reporte_time_planner.pdf')

def exportar_excel(datos):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.sheet_view.showGridLines = False  # quita las líneas de cuadrícula por defecto

    CAFE      = "3D1F0D"
    TERRACOTA = "C0623A"
    CREMA     = "FDF6EC"
    DORADO    = "8B5A2B"
    BLANCO    = "FFFFFF"

    borde_fino = Border(
        left=Side(style='thin', color="E8D5B0"),
        right=Side(style='thin', color="E8D5B0"),
        top=Side(style='thin', color="E8D5B0"),
        bottom=Side(style='thin', color="E8D5B0")
    )

    def titulo_seccion(fila, texto, col_fin, color_fondo=CAFE):
        celda = ws.cell(row=fila, column=1, value=texto)
        celda.font = Font(bold=True, color=BLANCO, size=12)
        celda.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=col_fin)
        ws.row_dimensions[fila].height = 22
        for c in range(1, col_fin + 1):
            ws.cell(row=fila, column=c).fill = PatternFill("solid", fgColor=color_fondo)
            ws.cell(row=fila, column=c).border = borde_fino

    def fila_encabezados(fila, encabezados):
        for j, h in enumerate(encabezados, start=1):
            c = ws.cell(row=fila, column=j, value=h)
            c.font = Font(bold=True, size=10, color=BLANCO)
            c.fill = PatternFill("solid", fgColor=DORADO)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = borde_fino
        ws.row_dimensions[fila].height = 20

    def fila_datos(fila, valores):
        for j, v in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=j, value=v)
            c.border = borde_fino
            c.alignment = Alignment(horizontal='center', vertical='center')
            if fila % 2 == 0:
                c.fill = PatternFill("solid", fgColor=CREMA)

    # Encabezado principal
    ws['A1'] = "TIME PLANNER — Reporte de Productividad"
    ws['A1'].font = Font(bold=True, size=16, color=CAFE)
    ws['A2'] = f"Semana del {datos['inicio_semana']} al {datos['fin_semana']}"
    ws['A2'].font = Font(size=11, color=DORADO)
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # Resumen general
    titulo_seccion(4, "RESUMEN GENERAL", 4)
    fila_encabezados(5, ["Tareas totales", "Completadas", "Pendientes", "Tiempo trabajado"])
    valores_resumen = [datos['total_tareas'], datos['completadas'], datos['pendientes'], datos['tiempo_total']]
    for j, v in enumerate(valores_resumen, start=1):
        c = ws.cell(row=6, column=j, value=v)
        c.font = Font(bold=True, size=13, color=CAFE)
        c.fill = PatternFill("solid", fgColor=CREMA)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = borde_fino
    ws.row_dimensions[6].height = 26

    # Tiempo por categoría
    fila = 9
    titulo_seccion(fila, "TIEMPO POR CATEGORÍA", 4, color_fondo=TERRACOTA)
    fila_encabezados(fila + 1, ["Categoría", "Tiempo", "% del total", "Tareas completadas"])
    for i, cat in enumerate(datos['por_categoria'], start=fila + 2):
        fila_datos(i, [cat['nombre'], cat['tiempo_formato'], f"{cat['porcentaje']}%", cat['completadas']])

    # Detalle de tareas de la semana
    fila_detalle = fila + max(len(datos['por_categoria']), 1) + 4
    etiquetas_estado = {'completada': 'Completada', 'pendiente': 'Pendiente'}
    titulo_seccion(fila_detalle, "DETALLE DE TAREAS DE LA SEMANA", 4)
    fila_encabezados(fila_detalle + 1, ["Tarea", "Estado", "Estimado", "Real"])
    for i, t in enumerate(datos['detalle_tareas_semana'], start=fila_detalle + 2):
        fila_datos(i, [t['titulo'], etiquetas_estado.get(t['estado'], t['estado']),
                       t['tiempo_estimado'], t['tiempo_real']])

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='reporte_time_planner.xlsx')